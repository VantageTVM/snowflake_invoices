import re
import io
from datetime import date
import pdfplumber
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font
import streamlit as st

st.set_page_config(page_title="Snowflake Usage Converter", layout="centered")
st.title("Snowflake Usage PDF to Excel")

FULL_MONTHS = ['January', 'February', 'March', 'April', 'May', 'June',
               'July', 'August', 'September', 'October', 'November', 'December']
SHORT_MONTHS = ['Jan', 'Feb', 'Mar', 'Apr', 'May', 'Jun',
                'Jul', 'Aug', 'Sep', 'Oct', 'Nov', 'Dec']
MONTHS = {m: i + 1 for i, m in enumerate(SHORT_MONTHS)}


def extract_rows(pdf_file):
    rows = []
    with pdfplumber.open(pdf_file) as pdf:
        for page in pdf.pages:
            for table in page.extract_tables():
                for row in table:
                    if not row or not row[0]:
                        continue
                    cell0 = str(row[0]).strip()
                    if re.match(r'^[A-Z][a-z]+-\d{4}$', cell0):
                        rows.append({
                            'month':    cell0,
                            'category': str(row[1]).strip() if len(row) > 1 and row[1] else '',
                            'units':    str(row[2]).strip() if len(row) > 2 and row[2] else '',
                            'usd':      str(row[3]).strip() if len(row) > 3 and row[3] else '',
                        })
    return rows


def parse_number(val):
    if not val or val.strip() in ('N/A', '-', '', 'None'):
        return None
    val = val.strip().replace(',', '')
    negative = val.startswith('(') and val.endswith(')')
    val = val.strip('()')
    try:
        num = float(val)
        return -num if negative else num
    except ValueError:
        return None


def parse_month(val):
    parts = val.split('-')
    if len(parts) == 2 and parts[0] in MONTHS:
        return date(int(parts[1]), MONTHS[parts[0]], 1)
    return val


def append_rows(ws, rows):
    ws.append(['USAGE MONTH', 'USAGE CATEGORY', 'UNITS CONSUMED', 'TOTAL USAGE (USD)'])
    for row in rows:
        is_total = row['units'].strip() == 'N/A' and row['category'].endswith('TOTAL')
        units_val = 'N/A' if is_total else parse_number(row['units'])
        usd_val   = parse_number(row['usd'])
        month_val = parse_month(row['month'])
        ws.append([month_val, row['category'], units_val, usd_val])
        r = ws.max_row
        ws.cell(r, 1).number_format = 'MMM-YY'
        if is_total:
            for col in range(1, 5):
                ws.cell(r, col).font = Font(bold=True)
        if units_val not in ('N/A', None):
            ws.cell(r, 3).number_format = '#,##0.000;(#,##0.000)'
        if usd_val is not None:
            ws.cell(r, 4).number_format = '#,##0.00;(#,##0.00)'


def build_excel(rows, existing_file=None):
    if existing_file:
        wb = load_workbook(existing_file)
        ws = wb.active
        ws.append([None, None, None, None])  # blank separator row
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = 'Monthly Usage'
    append_rows(ws, rows)
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def derive_filename(existing_name, new_month_str):
    """Replace the month name in the existing filename with the new month."""
    parts = new_month_str.split('-')
    new_short = parts[0]
    new_full = FULL_MONTHS[MONTHS[new_short] - 1]
    stem = existing_name.removesuffix('.xlsx')

    # Try full month names first (longer match takes priority)
    for old_full in sorted(FULL_MONTHS, key=len, reverse=True):
        m = re.search(old_full, stem, re.IGNORECASE)
        if m:
            return stem[:m.start()] + new_full + stem[m.end():] + '.xlsx'

    # Try abbreviated month names
    for old_short in SHORT_MONTHS:
        m = re.search(r'\b' + old_short + r'\b', stem, re.IGNORECASE)
        if m:
            return stem[:m.start()] + new_short + stem[m.end():] + '.xlsx'

    return stem + '.xlsx'


pdf_file   = st.file_uploader("Upload Snowflake usage PDF", type="pdf")
excel_file = st.file_uploader("Append to existing Excel (optional)", type=["xlsx"])

if pdf_file:
    with st.spinner("Processing..."):
        rows = extract_rows(pdf_file)
        excel_buf = build_excel(rows, existing_file=excel_file)

    action = "Appended" if excel_file else "Extracted"
    st.success(f"{action} {len(rows)} rows")

    new_month_str = rows[0]['month'] if rows else None
    if excel_file and new_month_str:
        default_name = derive_filename(excel_file.name, new_month_str)
    else:
        default_name = pdf_file.name.removesuffix('.pdf') + '.xlsx'

    custom_name = st.text_input("File name (optional)", placeholder=default_name)
    out_name = custom_name.strip() if custom_name.strip() else default_name
    if not out_name.endswith('.xlsx'):
        out_name += '.xlsx'

    st.download_button(
        label="Download Excel",
        data=excel_buf,
        file_name=out_name,
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
