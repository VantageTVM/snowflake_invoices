"""
Snowflake Usage Statement PDF -> Excel Converter
Usage: python snowflake_usage.py <input.pdf> [output.xlsx]
"""

import sys
import re
import pdfplumber
from openpyxl import Workbook
from openpyxl.styles import Font


def extract_rows(pdf_path):
    rows = []
    with pdfplumber.open(pdf_path) as pdf:
        for page in pdf.pages:
            for table in page.extract_tables():
                for row in table:
                    if not row or not row[0]:
                        continue
                    cell0 = str(row[0]).strip()
                    if re.match(r'^[A-Z][a-z]+-\d{4}$', cell0):
                        rows.append({
                            'month':    cell0,
                            'category': str(row[1]).strip() if len(row) > 1 and row[1] else '',
                            'units':    str(row[2]).strip() if len(row) > 2 and row[2] else '',
                            'usd':      str(row[3]).strip() if len(row) > 3 and row[3] else '',
                        })
    return rows


def parse_number(val):
    if not val or val.strip() in ('N/A', '-', '', 'None'):
        return None
    val = val.strip().replace(',', '')
    negative = val.startswith('(') and val.endswith(')')
    val = val.strip('()')
    try:
        num = float(val)
        return -num if negative else num
    except ValueError:
        return None


def write_excel(rows, output_path):
    wb = Workbook()
    ws = wb.active
    ws.title = 'Monthly Usage'

    ws.append(['USAGE MONTH', 'USAGE CATEGORY', 'UNITS CONSUMED', 'TOTAL USAGE (USD)'])

    for row in rows:
        is_total = row['units'].strip() == 'N/A' and row['category'].endswith('TOTAL')
        units_val = 'N/A' if is_total else parse_number(row['units'])
        usd_val   = parse_number(row['usd'])
        ws.append([row['month'], row['category'], units_val, usd_val])
        if is_total:
            for col in range(1, 5):
                ws.cell(ws.max_row, col).font = Font(bold=True)

    wb.save(output_path)
    print(f"Saved: {output_path}  |  {len(rows)} rows")


if __name__ == '__main__':
    if len(sys.argv) < 2:
        print("Usage: python snowflake_usage.py <input.pdf> [output.xlsx]")
        sys.exit(1)
    pdf_path = sys.argv[1]
    out_path = sys.argv[2] if len(sys.argv) > 2 else 'snowflake_usage.xlsx'
    rows = extract_rows(pdf_path)
    write_excel(rows, out_path)
